import argparse
import os
import re
import shutil
import sys

import openpyxl


def strip_lines(lines):
    """Remove empty lines and double spaces from the lines"""
    clean_lines = []
    # If lines is a string, create a list of one element
    if isinstance(lines, str):
        lines = [lines]
        return_list = False
    else:
        return_list = True
    for line in lines:
        line = re.sub(' +', ' ', line)
        line = line.strip()
        if line:
            clean_lines.append(line)

    if return_list:
        return clean_lines
    else:
        return clean_lines[0]


def get_column_data(sheet, column, return_none=False):
    if return_none:
        return [strip_lines(cell.value) for cell in sheet[column]]
    else:
        return [
            strip_lines(cell.value)
            for cell in sheet[column]
            if cell.value is not None
        ]


def get_dict(sheet, column1, column2, info_fase=2):
    keys = get_column_data(sheet, column1)[1:]
    values = get_column_data(sheet, column2)[1:]
    # If length of keys and values is different, return None
    if len(keys) != len(values):
        raise ValueError(
            f'Length of keys and values is different in Fase {info_fase}: {len(keys)} vs {len(values)}'
        )
    return dict(zip(keys, values))


def main(args):
    if '.' not in args.fases_filename:
        args.fases_filename += args.extension
        args.fases_filename = os.path.normpath(args.fases_filename)
    if '.' not in args.input_filename:
        args.input_filename += args.extension
    if '.' not in args.fase2_from_file and args.fase2_from_file != 'None':
        args.fase2_from_file += args.extension
    if '.' not in args.fase3_from_file and args.fase3_from_file != 'None':
        args.fase3_from_file += args.extension

    fases_filename = os.path.join(args.fases_filename)
    fases_workbook = openpyxl.load_workbook(filename=fases_filename)
    fases_sheet = fases_workbook.worksheets[0]

    # Load fase_n changes from columns
    keys = get_column_data(fases_sheet, args.fase2_1_col[0])[1:]
    values = get_column_data(fases_sheet, args.fase2_1_col[1])[1:]
    f2_1 = get_dict(
        fases_sheet,
        args.fase2_1_col[0],
        args.fase2_1_col[1],
        info_fase='fase 2.1',
    )
    f2_2 = get_dict(
        fases_sheet,
        args.fase2_2_col[0],
        args.fase2_2_col[1],
        info_fase='fase 2.2',
    )
    f2_3 = get_dict(
        fases_sheet,
        args.fase2_3_col[0],
        args.fase2_3_col[1],
        info_fase='fase 2.3',
    )
    f3 = get_dict(
        fases_sheet, args.fase3_col[0], args.fase3_col[1], info_fase='fase 3'
    )
    f4 = get_dict(
        fases_sheet, args.fase4_col[0], args.fase4_col[1], info_fase='fase 4'
    )

    f2 = f2_1.copy()
    f2.update(f2_2)
    f2.update(f2_3)

    from_file = [args.fase2_from_file, args.fase3_from_file, 'None']
    input_file = [
        args.input_filename,
        'fase_2' + args.extension
        if args.fase2_from_file == 'None'
        else args.fase2_from_file,
        'fase_3' + args.extension
        if args.fase3_from_file == 'None'
        else args.fase3_from_file,
    ]
    output_file = [
        'fase_2' + args.extension,
        'fase_3' + args.extension,
        'fase_4' + args.extension,
    ]

    f_dict = [f2, f3, f4]
    f_dict_changes = [0, 0, 0]
    n_words = [0, 0, 0]

    for i in range(len(from_file)):
        if from_file[i] == 'None':
            input_filename = os.path.join(input_file[i])
            print(f'Input file: {input_filename}')
            input_workbook = openpyxl.load_workbook(filename=input_filename)
            input_sheet = input_workbook.worksheets[0]

            # Output file = input file + editing
            output_filename = os.path.join(output_file[i])
            shutil.copy(input_filename, output_filename)
            output_workbook = openpyxl.load_workbook(filename=output_filename)
            output_sheet = output_workbook.worksheets[0]

            input_rows = get_column_data(
                input_sheet, args.input_col[0], return_none=True
            )[1:]

            for each_row in range(len(input_rows)):
                # print(f'Input row: {each_row}: {input_rows[each_row]}')
                if input_rows[each_row] is not None:
                    new_row = [
                        each[1:] if each[0] == ' ' else each
                        for each in input_rows[each_row].split(',')
                    ]
                    # print(f'After split: {new_row}')

                    n_words[i] += len(new_row)
                    for each_key in f_dict[i].keys():
                        for each_word in range(len(new_row)):
                            if new_row[each_word] == each_key:
                                f_dict_changes[i] += 1
                                # print(f'Replacing {new_row[each_word]} with {f_dict[i][each_key]}')
                                new_row[each_word] = f_dict[i][each_key]
                    # print(f'After replacement: {new_row}')

                    try:
                        new_row = ', '.join(new_row)
                    except Exception as e:
                        print(f'Error {e} in row {each_row}: {new_row}')
                        # print(new_row)

                else:
                    new_row = ''
                    print(f'Empty row: {each_row}: {new_row}')
                # print(f'Output row {each_row}: {new_row}\n')
                output_sheet[
                    args.input_col + str(each_row + 2)
                ].value = new_row
            output_workbook.save(filename=output_filename)

    print('\n Input files:')
    for each in input_file:
        print(' - {}'.format(each))

    print('\n Output files:')
    for each in output_file:
        print(' - {}'.format(each))

    print('\n {} words have been replaced in Fase 2'.format(f_dict_changes[0]))
    print(' {} words have been replaced in Fase 3'.format(f_dict_changes[1]))
    print(
        ' {} words have been replaced in Fase 4 \n'.format(f_dict_changes[2])
    )

    print(' {} items in Fase 2'.format(n_words[0]))
    print(' {} items in Fase 3'.format(n_words[1]))
    print(' {} items in Fase 4 \n'.format(n_words[2]))


if __name__ == '__main__':
    # Running the script as python ... is deprecated
    parser = argparse.ArgumentParser(
        description='Automated replacement program'
    )
    parser.add_argument(
        '-fases_filename',
        default='PRUEBA',
        help='Excel file with fases replacement columns',
    )
    parser.add_argument(
        '-input_filename',
        default='CA_UNIVERSIDAD_LEXICO_FASE1-2',
        help='Input file to replace words from',
    )
    parser.add_argument(
        '-fase2_1_col', default='AB', help='Name of two columns for fase 2.1'
    )
    parser.add_argument(
        '-fase2_2_col', default='CD', help='Name of two columns for fase 2.2'
    )
    parser.add_argument(
        '-fase2_3_col', default='EF', help='Name of two columns for fase 2.3'
    )
    parser.add_argument(
        '-fase3_col', default='GH', help='Name of two columns for fase 3'
    )
    parser.add_argument(
        '-fase4_col', default='JK', help='Name of two columns for fase 4'
    )
    parser.add_argument(
        '-input_col', default='D', help='Name of column for input file content'
    )
    parser.add_argument(
        '-fase2_from_file',
        default='None',
        help='Extract Fase 2 data from file',
    )
    parser.add_argument(
        '-fase3_from_file',
        default='None',
        help='Extract Fase 3 data from file',
    )

    args = parser.parse_args()
    main(args)
