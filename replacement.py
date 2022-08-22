import openpyxl
import argparse
import os
import shutil

parser = argparse.ArgumentParser(description='Automated replacement program')
parser.add_argument('-fases_filename', default='PRUEBA', help='Excel file with fases replacement columns')
parser.add_argument('-input_filename', default='CA_UNIVERSIDAD_LEXICO_FASE1-2', help='Input file to replace words from')
parser.add_argument('-extension', default='.xlsx', help='Default Excel files extension')
parser.add_argument('-fase2_1_col', default='AB', help='Name of two columns for fase 2.1')
parser.add_argument('-fase2_2_col', default='CD', help='Name of two columns for fase 2.2')
parser.add_argument('-fase2_3_col', default='EF', help='Name of two columns for fase 2.3')
parser.add_argument('-fase3_col', default='GH', help='Name of two columns for fase 3')
parser.add_argument('-fase4_col', default='JK', help='Name of two columns for fase 4')
parser.add_argument('-input_col', default='D', help='Name of column for input file content')
parser.add_argument('-fase2_from_file', default='None', help='Extract Fase 2 data from file')
parser.add_argument('-fase3_from_file', default='None', help='Extract Fase 3 data from file')

args = parser.parse_args()


def get_column_data(sheet, column):
    return [cell.value for cell in sheet[column] if cell.value is not None]

fases_filename = os.path.join('..', args.fases_filename + args.extension)
fases_workbook = openpyxl.load_workbook(filename=fases_filename)
fases_sheet = fases_workbook.worksheets[0]


# Load fase_n changes from columns
f2_1 = dict(zip(get_column_data(fases_sheet, args.fase2_1_col[0])[1:],
                get_column_data(fases_sheet, args.fase2_1_col[1])[1:]))
f2_2 = dict(zip(get_column_data(fases_sheet, args.fase2_2_col[0])[1:],
                get_column_data(fases_sheet, args.fase2_2_col[1])[1:]))
f2_3 = dict(zip(get_column_data(fases_sheet, args.fase2_3_col[0])[1:],
                get_column_data(fases_sheet, args.fase2_3_col[1])[1:]))
f3 = dict(zip(get_column_data(fases_sheet, args.fase3_col[0])[1:],
                get_column_data(fases_sheet, args.fase3_col[1])[1:]))
f4 = dict(zip(get_column_data(fases_sheet, args.fase4_col[0])[1:],
                get_column_data(fases_sheet, args.fase4_col[1])[1:]))

f2 = f2_1.copy()
f2.update(f2_2)
f2.update(f2_3)

from_file = [args.fase2_from_file, args.fase3_from_file, 'None']
input_file = [args.input_filename,
                'fase_2' if args.fase2_from_file == 'None' else args.fase2_from_file,
                'fase_3' if args.fase3_from_file == 'None' else args.fase3_from_file]
output_file = ['fase_2', 'fase_3', 'fase_4']

f_dict = [f2, f3, f4]
f_dict_changes = [0, 0, 0]

for i in range(len(from_file)):
    if from_file[i] == 'None':
        input_filename = os.path.join('..', input_file[i] + args.extension)
        input_workbook = openpyxl.load_workbook(filename=input_filename)
        input_sheet = input_workbook.worksheets[0]
        
        # Output file = input file + editing
        output_filename = os.path.join('..', output_file[i] + args.extension)
        shutil.copy(input_filename, output_filename)
        output_workbook = openpyxl.load_workbook(filename=output_filename)
        output_sheet = output_workbook.worksheets[0]

        input_rows = get_column_data(input_sheet, args.input_col[0])[1:]

        for each_row in range(len(input_rows)):
            new_row = [each[1:] if each[0] == ' ' else each for each in input_rows[each_row].split(',')]
            for each_key in f_dict[i].keys():
                for each_word in range(len(new_row)):
                    if new_row[each_word] == each_key:
                        f_dict_changes[i] += 1
                        new_row[each_word] = f_dict[i][each_key]

            try:
                new_row = ', '.join(new_row)
            except:
                print(new_row)
            output_sheet[args.input_col + str(each_row + 2)].value = new_row
        output_workbook.save(filename=output_filename)



print('\n Input files:')
for each in input_file:
    print(' - {}'.format(each))

print('\n Output files:')
for each in output_file:
    print(' - {}'.format(each))

print('\n {} words have been replaced in Fase 2'.format(f_dict_changes[0]))
print(' {} words have been replaced in Fase 3'.format(f_dict_changes[1]))
print(' {} words have been replaced in Fase 4 \n'.format(f_dict_changes[2]))




