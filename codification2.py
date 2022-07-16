import openpyxl
import argparse
import os
import shutil

parser = argparse.ArgumentParser(description='Automated replacement program')
parser.add_argument('-fases_filename', default='PRUEBA', help='Excel file with fases replacement columns')
parser.add_argument('-input_filename', default='CA_UNIVERSIDAD_LEXICO_FASE1-2', help='Input file to replace words from')
parser.add_argument('-extension', default='.xlsx', help='Default Excel files extension')
parser.add_argument('-fase2_1_col', default='AB', help='Name of two columns for fase 2.1')
parser.add_argument('-fase2_2_col', default='CD', help='Name of two columns for fase 2.2')
parser.add_argument('-fase2_3_col', default='EF', help='Name of two columns for fase 2.3')
parser.add_argument('-fase3_col', default='GH', help='Name of two columns for fase 3')
parser.add_argument('-fase4_col', default='JK', help='Name of two columns for fase 4')
parser.add_argument('-input_col', default='D', help='Name of column for input file content')

args = parser.parse_args()



def get_column_data(sheet, column):
    return [cell.value for cell in sheet[column] if cell.value is not None]

fases_filename = os.path.join('..', args.fases_filename + args.extension)
fases_workbook = openpyxl.load_workbook(filename=fases_filename)
fases_sheet = fases_workbook.worksheets[0]

input_filename = os.path.join('..', args.input_filename + args.extension)
input_workbook = openpyxl.load_workbook(filename=input_filename)
input_sheet = input_workbook.worksheets[0]

output_f2_filename = os.path.join('..', 'fase_2' + args.extension)
output_f3_filename = os.path.join('..', 'fase_3' + args.extension)
output_f4_filename = os.path.join('..', 'fase_4' + args.extension)
shutil.copy(input_filename, output_f2_filename)
shutil.copy(input_filename, output_f3_filename)
shutil.copy(input_filename, output_f4_filename)

output_f2_workbook = openpyxl.load_workbook(filename=output_f2_filename)
output_f2_sheet = output_f2_workbook.worksheets[0]
output_f3_workbook = openpyxl.load_workbook(filename=output_f3_filename)
output_f3_sheet = output_f3_workbook.worksheets[0]
output_f4_workbook = openpyxl.load_workbook(filename=output_f4_filename)
output_f4_sheet = output_f4_workbook.worksheets[0]


f2_1 = dict(zip(get_column_data(fases_sheet, args.fase2_1_col[0])[1:],
                get_column_data(fases_sheet, args.fase2_1_col[1])[1:]))
f2_2 = dict(zip(get_column_data(fases_sheet, args.fase2_2_col[0])[1:],
                get_column_data(fases_sheet, args.fase2_2_col[1])[1:]))
f2_3 = dict(zip(get_column_data(fases_sheet, args.fase2_3_col[0])[1:],
                get_column_data(fases_sheet, args.fase2_3_col[1])[1:]))
f3 = dict(zip(get_column_data(fases_sheet, args.fase3_col[0])[1:],
                get_column_data(fases_sheet, args.fase3_col[1])[1:]))
f4 = dict(zip(get_column_data(fases_sheet, args.fase4_col[0])[1:],
                get_column_data(fases_sheet, args.fase4_col[1])[1:]))


input_rows = get_column_data(input_sheet, args.input_col[0])[1:]
f2 = f2_1.copy()
f2.update(f2_2)
f2.update(f2_3)

f2_changes = 0
f3_changes = 0
f4_changes = 0

# Fase 2
for each_row in range(len(input_rows)):
    new_row = [each[1:] if each[0] == ' ' else each for each in input_rows[each_row].split(',')]
    for each_key in f2.keys():
        for each_word in range(len(new_row)):
            if new_row[each_word] == each_key:
                f2_changes += 1
                new_row[each_word] = f2[each_key]

    new_row = ', '.join(new_row)
    output_f2_sheet[args.input_col + str(each_row + 2)].value = new_row
output_f2_workbook.save(filename=output_f2_filename)

# Fase 3
input_rows = get_column_data(output_f2_sheet, args.input_col[0])[1:]
for each_row in range(len(input_rows)):
    new_row = [each[1:] if each[0] == ' ' else each for each in input_rows[each_row].split(',')]
    for each_key in f3.keys():
        for each_word in range(len(new_row)):
            if new_row[each_word] == each_key:
                f3_changes += 1
                new_row[each_word] = f3[each_key]

    new_row = ', '.join(new_row)
    output_f3_sheet[args.input_col + str(each_row + 2)].value = new_row
output_f3_workbook.save(filename=output_f3_filename)

# Fase 4
input_rows = get_column_data(output_f3_sheet, args.input_col[0])[1:]
for each_row in range(len(input_rows)):
    new_row = [each[1:] if each[0] == ' ' else each for each in input_rows[each_row].split(',')]
    for each_key in f4.keys():
        for each_word in range(len(new_row)):
            if new_row[each_word] == each_key:
                f4_changes += 1
                new_row[each_word] = f4[each_key]

    new_row = ', '.join(new_row)
    output_f4_sheet[args.input_col + str(each_row + 2)].value = new_row
output_f4_workbook.save(filename=output_f4_filename)


print('\n Input file: {}'.format(input_filename))
print(' Output files:')
print(' - {}'.format(output_f2_filename))
print(' - {}'.format(output_f3_filename))
print(' - {}'.format(output_f4_filename))
print('\n {} words have been replaced in Fase 2'.format(f2_changes))
print(' {} words have been replaced in Fase 3'.format(f3_changes))
print(' {} words have been replaced in Fase 4 \n'.format(f4_changes))




